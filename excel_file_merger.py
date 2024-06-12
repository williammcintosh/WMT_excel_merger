import openpyxl
import datetime
import os
import calendar
from tqdm import tqdm

# Declare log_file as a global variable
log_file = None

def log_message(message):
    global log_file
    with open(log_file, 'a') as f:
        f.write(message + '\n')

def get_valid_year():
    while True:
        input_str = input("Enter the year\t: ")
        try:
            # Convert the input to an integer
            input_year = int(input_str)
            
            # Check if the input is a two-digit year
            if 0 <= input_year <= 99:
                return input_year
            
            # Check if the input is a four-digit year
            if 2022 <= input_year <= 2099:
                return input_year - 2000
            
            # If the input is out of expected range
            print("Invalid year. Please enter a year between 2022 and 2099, or a two-digit year.")
        except ValueError:
            print("Invalid input. Please enter a valid integer.")


def get_valid_month():
    months = [month.lower() for month in calendar.month_name[1:]]  # Lowercase month names for easier comparison
    while True:
        input_month = input("Enter the month\t: ").strip().lower()
        matches = [month for month in months if month.startswith(input_month)]
        if len(matches) == 1:
            return matches[0].capitalize()
        elif len(matches) > 1:
            print(f"Ambiguous input. Did you mean one of these: {', '.join(matches)}?")
        else:
            print("Invalid month. Please enter a valid month name or partial name.")


def log_mapping_row(map_row):
    log_message(f"""
    {'Port_site:':<30} | {'mastersheet_sheet':<30} | {'mastersheet_material':<30}
    {'-'*90}
    {map_row[0].value:<30} | {map_row[1].value:<30} | {map_row[2].value:<30}
    """)

# This function is for Location ref# 1207905-4.39
def update_mastersheet_material_type(notes):
    hazwords = ["batter", "hid", "e waste", "light", "tube", "bulb"]
    # Default to "metal", "wood"
    mat = "Misc."
    if notes:
        if "film" in notes.lower():
            mat = "Plastics (i.e, film, rigids)"
        elif "glass" in notes.lower():
            mat = "Glass"
        elif "grease" in notes.lower():
            mat =  "Grease ORCO / PPV"
        elif any(word in notes.lower() for word in hazwords):
            mat = "Universal/Hazardous Waste"
    return "".join(mat.split()).lower()


def get_port_waste_sheet(input_month, input_year, port_waste_file):
    input_year += 2000
    all_months = list(calendar.month_name)
    first_half = all_months[7:13]
    port_waste_tabs = port_waste_file.sheetnames

    if input_month in first_half:
        tab = f'{input_year}-{input_year+1}'
    else:
        tab = f'{input_year-1}-{input_year}'
    
    log_message(f"Selected tab: {tab}")
    port_waste_tab = tab if tab in port_waste_tabs else None
    if port_waste_tab:
        return port_waste_file[port_waste_tab]
    return None


def get_formatted_datetime():
    current_datetime = datetime.datetime.now()
    formatted_datetime = current_datetime.strftime("%Y-%m-%d__%H-%M-%S")
    return formatted_datetime


def setup_logging(formatted_datetime):
    global log_file
    log_folder = "Logs"
    os.makedirs(log_folder, exist_ok=True)
    log_file = os.path.join(log_folder, f"Log_{formatted_datetime}.txt")
    log_message("Logging started.")


def get_port_waste_row(map_row, port_waste_sheet, location):
    log_message(f"""
        Locating Row from Port Waste:
                        {'FOUND VAL':<30} | {'DESIRED VAL':<30}
        {'-'*60}
    """)
    for row in port_waste_sheet.iter_rows(min_row=4):
        port_loc = (row[0].value or "BLANK_CELL").replace('\n', ' ')
        map_loc = (map_row[0].value or "BLANK_CELL").replace('\n', ' ')
        log_message(f"\t\t\t{port_loc:<30} | {map_loc:<30}")
        try:
            port_location = "".join(port_loc.split()).lower()
        except:
            pass
        if port_location == location: 
            return row[1].row
    return None


def get_port_waste_column(port_waste_sheet, input_month):
    log_message(f"""
        Locating Column from Port Waste:
                        {'FOUND VAL':<30} | {'DESIRED VAL':<30}
        {'-'*60}
    """)
    # Find the column in port_waste.xlsx with matching input_month
    for col in port_waste_sheet.iter_cols(min_col=10):
        log_message(f"\t\t\t{col[1].value:<30} | {input_month:<30}")
        if col[1].value == input_month:
            return col[1].column
    return None


def get_port_waste_cell_n_value_n_note(port_waste_sheet, port_waste_row, port_waste_column, map_row):
    if port_waste_row is not None and port_waste_column is not None:
        # A match was found, store the cell in port_waste.xlsx
        port_waste_cell = port_waste_sheet.cell(row=port_waste_row, column=port_waste_column)
        port_waste_value = port_waste_cell.internal_value
        port_waste_note = port_waste_cell.comment.text if port_waste_cell.comment else None
    else:
        if port_waste_row is None:
            log_message(f'''
                ERROR: Could not find the desired row in the 'PORT WASTE COLLECTION  RECY REPORT.xlsx' file.
                Looking for:
                \t'{map_row[0].value}' under "Location"
                \t'{map_row[1].value}' under "Site Description"
                \t'{map_row[2].value}' under "Material Collected"
                Please check the spelling and values in both the 'Mapping.xlsx' and 'PORT WASTE COLLECTION  RECY REPORT.xlsx' files.
            ''')
        if port_waste_column is None:
            log_message(f'''
                ERROR: Compare the spelling of '{input_month}' with the columns of the 'PORT WASTE COLLECTION  RECY REPORT.xlsx' file!
            ''')
    # Reset value to zero if it's blank
    port_waste_cell = 0 if port_waste_cell is None else port_waste_cell
    # return all three
    return port_waste_cell, port_waste_value, port_waste_note


def get_mastersheet_material_type(map_row, port_waste_note):
    mastersheet_material_type = "".join(map_row[2].value.split()).lower()
    if mastersheet_material_type == "checkcellnotes":
            mastersheet_material_type = update_mastersheet_material_type(port_waste_note)
    return mastersheet_material_type


def get_mastersheet_column(mastersheet_sheet,mastersheet_material_type):
    log_message(f"""
        Locating Column from Master:
                        {'FOUND VAL':<30} | {'DESIRED VAL':<30}
        {'-'*60}
    """)
    for col in mastersheet_sheet.iter_cols():
        mastersheet_mat = col[2].value or "BLANK_CELL"
        log_message(f"\t\t\t{mastersheet_mat:<30} | {mastersheet_material_type:<30}")
        if col[2].value is not None and "".join(mastersheet_mat.split()).lower() == mastersheet_material_type:
            return col[2]
    return None 


def get_mastersheet_row(mastersheet_sheet, input_year):
    log_message(f"""
        Locating Row from Master:
                        {'FOUND VAL':<30} | {'DESIRED VAL':<30}
        {'-'*60}
    """)
    # Find the row that matches the date
    for row in mastersheet_sheet.iter_rows(min_row=3):
        input_date = f"{input_month[:3]}-{input_year}"
        mastersheet_date = ""
        try:
            date_obj = datetime.datetime.strptime(str(row[0].value), "%Y-%m-%d %H:%M:%S")
            year = date_obj.strftime("%y")
            month = date_obj.strftime("%b")
            mastersheet_date = f"{month}-{year}"
        except:
            pass

        log_message(f"\t\t\t{mastersheet_date:<30} | {input_date:<30}")
        
        if mastersheet_date == input_date:
            return row[4]
    return None


def get_mastersheet_cell(mastersheet_sheet, mastersheet_row, mastersheet_column, map_row):
    if mastersheet_row is not None and mastersheet_column is not None:
        return mastersheet_sheet.cell(row=mastersheet_row.row, column=mastersheet_column.column)
    else:
        if mastersheet_column is None and mastersheet_row is not None:
            log_message(f'''
                Could not find the desired column in the 'Mastersheet.xlsx' file.
                Looking for:
                \t '{map_row[3].value}' under "Tab"
                \t'{map_row[4].value}' under "Type"
                Please check the spelling and values in both the 'Mapping.xlsx' and 'Mastersheet.xlsx' files.
            ''')
        elif mastersheet_row is None and mastersheet_column is not None:
            log_message(f'''
                Compare the spelling of '{input_month}' with the rows of the 'Mastersheet.xlsx' file!
            ''')
    return None


def update_mastersheet_cell_value_n_comment(mastersheet_cell, port_waste_value, port_waste_note):
    # Update the formula in the master cell
    if mastersheet_cell is not None and port_waste_value is not None:
        if mastersheet_cell.value is not None:
            formula = f"{mastersheet_cell.internal_value}+{port_waste_value}"
        else:
            formula = f"=0+{port_waste_value}"
        mastersheet_cell.value = formula
        if port_waste_note is not None:
            mastersheet_cell.comment = openpyxl.comments.Comment(port_waste_note, "Author")
    

def excel_merger(input_year, input_month):
    
    formatted_datetime = get_formatted_datetime()

    setup_logging(formatted_datetime)

    # Attempt to open the `PORT WASTE COLLECTION  RECY REPORT.xlsx` file
    try:
        port_waste_file = openpyxl.load_workbook('PORT WASTE COLLECTION  RECY REPORT.xlsx', data_only=True)
    except openpyxl.utils.exceptions.InvalidFileException as e:
        if type(e) == openpyxl.utils.exceptions.InvalidFileException:
            log_message('''
                ERROR -
                Open 'PORT WASTE COLLECTION  RECY REPORT.xls' in Excel and save as
                'PORT WASTE COLLECTION  RECY REPORT.xlsx' with the 'xlsx' ending,
                then try again. This program doesn't accept 'xls' endings.
            ''')
        else:
            log_message('An unknown error occurred.')
        return

    # Get the correct tab from the port_waste file
    port_waste_sheet = get_port_waste_sheet(input_month, input_year, port_waste_file)   

    # Open the mapping.xlsx file
    mapping_file = openpyxl.load_workbook('mapping.xlsx')
    mapping_sheet = mapping_file.active

    # Open the master.xlsx file
    mastersheet_file = openpyxl.load_workbook('Mastersheets.xlsx')

    # Get the total number of rows, excluding the header row
    total_rows = len(list(mapping_sheet.iter_rows(min_row=2)))-2

    for map_row in tqdm(mapping_sheet.iter_rows(min_row=2), desc="Processing rows", total=total_rows):
        # Get values for each row from Mapping file
        try:
            port_waste_location = "".join(map_row[0].value.split()).lower()
            mastersheet_tab = map_row[1].value
            port_waste_note = None
        except:
            break  # End Of File, break out!

        log_mapping_row(map_row)

        port_waste_row = get_port_waste_row(map_row, port_waste_sheet, port_waste_location)

        port_waste_column = get_port_waste_column(port_waste_sheet, input_month)

        port_waste_cell, port_waste_value, port_waste_note = get_port_waste_cell_n_value_n_note(port_waste_sheet, port_waste_row, port_waste_column, map_row)

        mastersheet_sheet = mastersheet_file[mastersheet_tab]

        mastersheet_material_type = get_mastersheet_material_type(map_row, port_waste_note)
        
        mastersheet_column = get_mastersheet_column(mastersheet_sheet, mastersheet_material_type)
        
        mastersheet_row = get_mastersheet_row(mastersheet_sheet, input_year)

        mastersheet_cell = get_mastersheet_cell(mastersheet_sheet, mastersheet_row, mastersheet_column, map_row)

        update_mastersheet_cell_value_n_comment(mastersheet_cell, port_waste_value, port_waste_note)

        log_message(f"""
            port_waste_cell\t: {port_waste_cell}
            port_waste_value\t: {port_waste_value or 'BLANK_CELL'}
            port_waste_note\t: {port_waste_note or 'NO_NOTE'}
            mastersheet_cell\t: {mastersheet_cell}
            \n
        """)

    # Save the changes to master.xlsx
    new_file_name = f"Mastersheets_{formatted_datetime}.xlsx"
    mastersheet_file.save(new_file_name)
    log_message(f"Master file saved as {new_file_name}")
    log_message("Logging ended.")


if __name__ == "__main__":
    # Get user input for year and month
    input_year = get_valid_year()
    input_month = get_valid_month()
    # Run main program
    excel_merger(input_year, input_month)